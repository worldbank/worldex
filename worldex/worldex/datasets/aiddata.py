"""
Automates indexing of aiddata pop datasets
"""
import csv
import os
from datetime import date, datetime
from pathlib import Path
from unicodedata import normalize

import pandas as pd
import requests
from bs4 import BeautifulSoup
from pyunpack import Archive
from shapely import wkt
from shapely.geometry import box
from shapely.ops import unary_union
from openpyxl import load_workbook

from ..handlers.vector_handlers import (
    VectorHandler,
    POSSIBLE_GEOM,
    POSSIBLE_X,
    POSSIBLE_Y,
)
from ..utils.filemanager import download_file
from .dataset import BaseDataset


class AidDataDataset(BaseDataset):
    source_org: str = "AidData"

    @classmethod
    def from_url(cls, url: str):
        """
        >>> url = "https://www.aiddata.org/data/korea-koica-project-database-2-2009"
        >>> AidDataDataset.from_url(url)
        """
        response = requests.get(url)
        # Aiddata encoding is weird.
        # To fix this, we set encoding to apparent encoding and normalize rendered texts
        # TODO: support unary union
        response.encoding = response.apparent_encoding
        soup = BeautifulSoup(response.text, "html.parser")
        title = normalize("NFKC", soup.find("h2").text)
        description = normalize(
            "NFKC", soup.find(string="Summary").parent.find_next().text
        )
        file = soup.find(string="Download").parent["href"]
        return cls(
            name=title,
            files=[file],
            last_fetched=datetime.now().isoformat(),
            data_format="CSV",
            description=description,
            projection="EPSG:4326",
            properties={},
            keywords=[],
            accessibility="public/open",
            url=url,
        )

    def download(self):
        """Download all files"""
        for file in self.files:
            filename = Path(file).name
            if not os.path.exists(self.dir / filename):
                # TODO: https download is way slower than using worldpop ftp
                download_file(file, self.dir / filename)

    def unzip(self):
        """Unzip all files"""
        for file in filter(
            lambda x: x.endswith(".zip") or x.endswith(".7z"), self.files
        ):
            filename = Path(file).name
            Archive(self.dir / filename).extractall(self.dir)

    @staticmethod
    def is_valid_csv(csv_file):
        with open(csv_file, "r") as f:
            dict_reader = csv.DictReader(f)
            headers = [h.lower() for h in dict_reader.fieldnames]
            return (
                any(p_x in headers for p_x in POSSIBLE_X)
                and any(p_y in headers for p_y in POSSIBLE_Y)
            ) or any(p_g in headers for p_g in POSSIBLE_GEOM)

    @staticmethod
    def is_valid_excel(excel_file):
        worksheets = load_workbook(excel_file)
        columns = []
        for sheet in worksheets:
            for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                columns.append(value.lower())
        return (
            any(p_x in columns for p_x in POSSIBLE_X)
            and any(p_y in columns for p_y in POSSIBLE_Y)
        ) or any(p_g in columns for p_g in POSSIBLE_GEOM)

    def index(self, window=(10, 10)):
        self.download()
        self.unzip()
        boxes = []
        indices = []
        csv_files = self.dir.glob("**/*.csv")
        for file in csv_files:
            if not self.is_valid_csv(file):
                continue
            handler = VectorHandler.from_csv(file)
            h3indices = handler.h3index()
            boxes.append(box(*handler.bbox))
            indices.append(pd.DataFrame({"h3_index": h3indices}))
        xlsx_files = self.dir.glob("**/*.xlsx")
        for file in xlsx_files:
            if not self.is_valid_excel(file):
                continue
            handler = VectorHandler.from_excel(file)
            h3indices = handler.h3index()
            boxes.append(box(*handler.bbox))
            indices.append(pd.DataFrame({"h3_index": h3indices}))
        # TODO: raise if not valid file
        self.bbox = wkt.dumps(box(*unary_union(boxes).bounds))
        df = pd.concat(indices).drop_duplicates()
        self.write(df)
        return df
